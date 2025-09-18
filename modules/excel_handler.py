from pathlib import Path
from openpyxl import load_workbook, Workbook

def extract_responses(excel_path):
    """Extrai todas as respostas de uma planilha Excel para uma lista de dicionários."""
    wb = load_workbook(excel_path)
    ws = wb.active
    
    headers = [cell.value.strip() if cell.value else "" for cell in ws[1]]
    responses = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # Ignora linhas vazias
            continue
        response_dict = {header: value for header, value in zip(headers, row)}
        responses.append(response_dict)
    
    return responses

def filter_new_responses(validation_file_path, all_responses, id_column_name='ID'):
    """Filtra apenas as respostas que ainda não foram registradas."""
    validation_path = Path(validation_file_path)
    validation_path.parent.mkdir(parents=True, exist_ok=True)
    
    registered_ids = set()
    if validation_path.exists():
        wb_val = load_workbook(validation_path)
        ws_val = wb_val.active
        # Encontra o índice da coluna de ID
        try:
            headers = [cell.value for cell in ws_val[1]]
            id_index = headers.index(id_column_name)
            for row in ws_val.iter_rows(min_row=2, values_only=True):
                if row[id_index] is not None:
                    registered_ids.add(str(row[id_index]))
        except (ValueError, IndexError):
            print(f"Aviso: Coluna '{id_column_name}' não encontrada em {validation_file_path}.")

    new_responses = [resp for resp in all_responses if str(resp.get(id_column_name)) not in registered_ids]
    return new_responses

def update_validation_file(validation_file_path, new_responses):
    """Adiciona as novas respostas ao arquivo de validação."""
    if not new_responses:
        return
        
    validation_path = Path(validation_file_path)
    
    if not validation_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.append(list(new_responses[0].keys())) # Cria cabeçalho
    else:
        wb = load_workbook(validation_path)
        ws = wb.active
        
    for response in new_responses:
        # Garante que a ordem das colunas seja a mesma do cabeçalho
        headers = [cell.value for cell in ws[1]]
        row_to_append = [response.get(header, "") for header in headers]
        ws.append(row_to_append)
        
    wb.save(validation_path)